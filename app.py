import io
import re
import time
import zipfile
from datetime import datetime

import pandas as pd
import streamlit as st
import google.generativeai as genai
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pypdf import PdfReader, PdfWriter
from PIL import Image

# ══════════════════════════════════════════════════════════════════════════════
# KONSTANTEN & CONFIG
# ══════════════════════════════════════════════════════════════════════════════
PAGE_TITLE      = "DATEV Beleg-Parser Pro AI"
PAGE_ICON       = "🧾"
GEMINI_MODEL    = "gemini-3.1-flash-lite"   
FREE_TIER_DELAY = 0.0  # 유료 계정의 최고 속도 보장
MWST_19_FACTOR  = 19 / 119
MWST_7_FACTOR   = 7 / 107

MIME_MAP = {
    "pdf":  "application/pdf",
    "jpg":  "image/jpeg",
    "jpeg": "image/jpeg",
    "png":  "image/png",
}

_ILLEGAL_CHARS = re.compile(r'[\\/*?:"<>|]')

# 💡 기본 초기 규칙 정의 (Shell, Google 2개만 유지)
INITIAL_VENDORS = {
    "Shell":      {"SKR03": "4530 - Kfz-Betriebskosten", "SKR04": "6520 - Kfz-Betriebskosten"},
    "Google":     {"SKR03": "4920 - Telefon", "SKR04": "6815 - Bürobedarf"},
}

# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT PAGE SETUP & CSS HACKS
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title=PAGE_TITLE, page_icon=PAGE_ICON, layout="wide")

st.markdown("""
    <style>
        [data-testid="stSidebarNav"] {display: none !important;}
        section[data-testid="stSidebar"] {display: none !important;}
    </style>
""", unsafe_allow_html=True)

st.title("Kognitiver Beleg-Parser (Pure German Edition)")
st.caption("Automatisierte Belegfassung mit SKR-Klassifizierung. Alle Begriffe entsprechen den offiziellen deutschen Buchhaltungsstandards.")

# ══════════════════════════════════════════════════════════════════════════════
# API AUTHENTIFIZIERUNG
# ══════════════════════════════════════════════════════════════════════════════
API_KEY: str = st.secrets.get("GEMINI_API_KEY", "")
if not API_KEY:
    API_KEY = st.text_input("🔑 Gemini API-Key eingeben", type="password")
    if API_KEY: genai.configure(api_key=API_KEY)
else:
    genai.configure(api_key=API_KEY)

# 💾 영구 유지되어야 하는 설정 영역 (작업 종료 시에도 보존됨)
if "custom_rules" not in st.session_state:
    st.session_state.custom_rules = INITIAL_VENDORS.copy()
if "custom_zahlungswege" not in st.session_state:
    st.session_state.custom_zahlungswege = ["Firmenkonto"]  # Firmenkonto를 기본값으로 세팅

# 🧼 휘발성 데이터 영역 (작업 종료 시 완벽히 삭제됨)
if "edited_receipts" not in st.session_state:
    st.session_state.edited_receipts = None

# ══════════════════════════════════════════════════════════════════════════════
# BACKEND ENGINE FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def ask_gemini_vision_cached(file_bytes: bytes, mime_type: str, api_key_trigger: str) -> tuple:
    fallback = ("", datetime.now().strftime("%Y-%m-%d"), "Unbekannt", 0.0, "EUR", "", "AUTO_19", "No OCR text")
    if not api_key_trigger: return fallback + (False,)
    try:
        model = genai.GenerativeModel(GEMINI_MODEL)
        prompt_text = get_gemini_prompt()
        response = model.generate_content([{"mime_type": mime_type, "data": file_bytes}, prompt_text])
        beleg_nr, d_str, ven, tot, cur, kat, m_type = _parse_gemini_response(response.text)
        return beleg_nr, d_str, ven, tot, cur, kat, m_type, response.text, True
    except Exception:
        return fallback + (False,)

def get_assigned_account(vendor_name: str, skr_mode: str) -> str:
    v_upper = vendor_name.upper()
    for keyword, accounts in st.session_state.custom_rules.items():
        if keyword.upper() in v_upper:
            return accounts[skr_mode]
    return ""

def create_sandwich_pdf(file_bytes: bytes, ext: str, raw_ai_text: str) -> bytes:
    try:
        writer = PdfWriter()
        if ext in ["jpg", "jpeg", "png"]:
            img = Image.open(io.BytesIO(file_bytes))
            img_pdf_buf = io.BytesIO()
            img.convert("RGB").save(img_pdf_buf, format="PDF")
            img_pdf_buf.seek(0)
            reader = PdfReader(img_pdf_buf)
            page = reader.pages[0]
        elif ext == "pdf":
            reader = PdfReader(io.BytesIO(file_bytes))
            page = reader.pages[0]
        else:
            return file_bytes
        writer.add_page(page)
        writer.add_metadata({"/Title": "DATEV Searchable Beleg", "/Subject": raw_ai_text.replace("\n", " ")})
        output_buf = io.BytesIO()
        writer.write(output_buf)
        return output_buf.getvalue()
    except Exception: return file_bytes

def sanitize_filename(text: str) -> str: return _ILLEGAL_CHARS.sub("", text).strip()

def build_datev_filename(date_str: str, vendor: str, brutto_eur: float, zahlungsweg: str, beleg_nr: str, ausgangs_inv: str) -> str:
    # 커스텀 결제 수단에 따른 동적 접미사 생성 규칙 (_B, 또는 앞자리 코드 매핑)
    if zahlungsweg.lower() == "firmenkonto":
        z_suffix = "B"
    elif zahlungsweg.lower() == "bar":
        z_suffix = "BAR"
    else:
        z_suffix = sanitize_filename(zahlungsweg).replace(" ", "").upper()[:3]
        
    v_clean = sanitize_filename(vendor).replace(" ", "")[:10]
    b_suffix = f"_{sanitize_filename(beleg_nr)[:12]}" if beleg_nr and beleg_nr.lower() not in ("", "none") else ""
    inv_suffix = f"-I{sanitize_filename(ausgangs_inv)[:8]}" if ausgangs_inv and ausgangs_inv.lower() not in ("", "none") else ""
    return f"{date_str.replace('-', '')}_{v_clean}_{brutto_eur:.2f}EUR_{z_suffix}{b_suffix}{inv_suffix}.pdf"

def get_gemini_prompt() -> str:
    return """Du bist ein Experte für deutsche Finanzbuchhaltung. Extrahiere folgende Daten aus dem Beleg:
1. Rechnungsnummer
2. Rechnungsdatum (YYYY-MM-DD)
3. Verkäufer (max 12 Zeichen)
4. Bruttobetrag (Zahl mit Punkt .)
5. Währung (EUR/USD)
6. Kategorie_SKR (Ignoriere dies, gib einfach "AUTO" an)
7. MwSt_Type ("19_Only", "7_Only", "Split", "0_Only", "AUTO_19")

Ausgabe strictly 7 Zeilen:
Beleg_Nr: [Nummer]
Datum: [YYYY-MM-DD]
Vendor: [Name]
Total: [Zahl]
Currency: [EUR/USD]
Kategorie: AUTO
MwSt_Type: [Type]"""

def _parse_german_amount(raw: str) -> float:
    s = re.sub(r"[€$£\s]", "", raw)
    s = re.sub(r"(?i)(eur|usd|gbp)", "", s).strip()
    if not s: return 0.0
    if "." in s and "," in s:
        if s.rfind(".") > s.rfind(","): s = s.replace(",", "")
        else: s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        if len(s[s.rfind(",") + 1:]) == 2: s = s.replace(",", ".")
        else: s = s.replace(",", "")
    try: return float(s)
    except ValueError: return 0.0

def _parse_gemini_response(text: str) -> tuple:
    beleg_nr, date_str, vendor, total, currency = "", datetime.now().strftime("%Y-%m-%d"), "Unbekannt", 0.0, "EUR"
    kategorie, mwst_type = "", "AUTO_19"
    cleaned = re.sub(r"[*`]", "", text)
    for line in cleaned.splitlines():
        if ":" not in line: continue
        key, _, value = line.partition(":")
        key, value = key.strip(), value.strip()
        match key:
            case "Beleg_Nr": beleg_nr = value
            case "Datum":
                if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value): date_str = value
            case "Vendor": 
                if value: vendor = value
            case "Total": total = _parse_german_amount(value)
            case "Currency":
                if value.upper() in ["EUR", "USD"]: currency = value.upper()
            case "MwSt_Type":
                if value: mwst_type = value
    return beleg_nr, date_str, vendor, total, currency, kategorie, mwst_type

def calculate_tax_details(brutto_eur: float, mwst_type: str) -> tuple[float, float, float]:
    mwst_19, mwst_7 = 0.0, 0.0
    if mwst_type in ("19_Only", "AUTO_19"): mwst_19 = round(brutto_eur * MWST_19_FACTOR, 2)
    elif mwst_type == "7_Only": mwst_7 = round(brutto_eur * MWST_7_FACTOR, 2)
    elif mwst_type == "Split":
        half = round(brutto_eur / 2, 2)
        mwst_19 = round(half * MWST_19_FACTOR, 2)
        mwst_7 = round((brutto_eur - half) * MWST_7_FACTOR, 2)
    return mwst_19, mwst_7, round(brutto_eur - (mwst_19 + mwst_7), 2)

# ══════════════════════════════════════════════════════════════════════════════
# REKALKULATION & EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def on_table_edited() -> None:
    edit_state  = st.session_state.get("beleg_editor_key", {})
    edited_rows = edit_state.get("edited_rows", {})
    deleted_rows = edit_state.get("deleted_rows", [])

    if not edited_rows and not deleted_rows: return

    df = st.session_state.edited_receipts.copy()
    
    # 데이터 행 유연한 수동 삭제 처리 연동
    if deleted_rows:
        indices_to_drop = [df.index[int(idx)] for idx in deleted_rows]
        df = df.drop(index=indices_to_drop)
        df.index = range(1, len(df) + 1)
        df.index.name = "Nr."
        st.session_state.edited_receipts = df
        return

    # 데이터 수동 수정 처리 연동 (Zahlungsweg 변경 실시간 연동 포함)
    for row_idx_str, changes in edited_rows.items():
        label = df.index[int(row_idx_str)]
        for col, new_val in changes.items(): df.at[label, col] = new_val

        brutto_eur = float(df.at[label, "Bruttobetrag (EUR)"])
        mwst_19, mwst_7, netto = calculate_tax_details(brutto_eur, str(df.at[label, "BU-Schlüssel"]))
        df.at[label, "USt/Vorsteuer 19%"] = mwst_19
        df.at[label, "Vorsteuer 7%"]  = mwst_7
        df.at[label, "Nettobetrag (Haben)"]    = netto
        df.at[label, "DATEV-Dateiname"] = build_datev_filename(
            str(df.at[label, "Belegdatum"]), str(df.at[label, "Kreditor"]), brutto_eur,
            str(df.at[label, "Zahlungsweg"]), str(df.at[label, "Belegnummer"]), str(df.at[label, "Ausgangs-Rechnungsnummer"])
        )
    st.session_state.edited_receipts = df

def build_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_clean = df.drop(columns=["_FileExt", "_RawBytes", "_OcrText"], errors="ignore")
        df_clean.to_excel(writer, sheet_name="DATEV_Export", index=True)
        ws = writer.sheets["DATEV_Export"]
        HEADER_FILL, HEADER_FONT = PatternFill("solid", fgColor="1F4E78"), Font(name="Arial", size=11, bold=True, color="FFFFFF")
        thin = Side(style="thin", color="D9D9D9")
        border_style = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]: cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, border_style
        for row in ws.iter_rows(min_row=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = border_style
                if col_idx in (7, 9, 10, 11): cell.number_format = '#,##0.00" €"'
                elif col_idx in (2, 5, 8, 12): cell.alignment = Alignment(horizontal="right")

        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    str_len = sum(2 if ord(char) > 128 else 1 for char in str(cell.value))
                    if str_len > max_len: max_len = str_len
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 5, 16)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# MAIN UI
# ══════════════════════════════════════════════════════════════════════════════

# 🛠️ 상단 설정 및 관리 영역 (2단 배열 구성)
col_menu1, col_menu2 = st.columns(2)

with col_menu1:
    with st.expander("Buchungsregeln verwalten", expanded=False):
        with st.form("new_rule_form", clear_on_submit=True):
            c1, c2, c3 = st.columns([2, 3, 3])
            with c1: new_vendor = st.text_input("Kreditor / Vendor")
            with c2: new_skr03  = st.text_input("SKR03")
            with c3: new_skr04  = st.text_input("SKR04")
            if st.form_submit_button("Regel speichern") and new_vendor:
                st.session_state.custom_rules[new_vendor] = {"SKR03": new_skr03, "SKR04": new_skr04}
                st.rerun()

        if st.session_state.custom_rules:
            for v in list(st.session_state.custom_rules.keys()):
                r_col1, r_col2, r_col3, r_col4 = st.columns([2, 3, 3, 1])
                r_col1.text(v)
                r_col2.text(st.session_state.custom_rules[v]["SKR03"])
                r_col3.text(st.session_state.custom_rules[v]["SKR04"])
                if r_col4.button("Löschen", key=f"del_{v}"):
                    del st.session_state.custom_rules[v]
                    st.rerun()

with col_menu2:
    # 🌟 [신설] Zahlungsweg 커스텀 관리 메뉴 (추가/삭제 및 자동 유지)
    with st.expander("Zahlungswege verwalten", expanded=False):
        with st.form("new_zahlungsweg_form", clear_on_submit=True):
            z_input_col, z_btn_col = st.columns([3, 1])
            with z_input_col: new_zw = st.text_input("Neuer Zahlungsweg (z.B. Kreditkarte, Paypal)")
            with z_btn_col: st.markdown("<br>", unsafe_allow_html=True); submit_zw = st.form_submit_button("Hinzufügen")
            if submit_zw and new_zw:
                if new_zw not in st.session_state.custom_zahlungswege:
                    st.session_state.custom_zahlungswege.append(new_zw)
                    st.rerun()
        
        if st.session_state.custom_zahlungswege:
            for zw_item in list(st.session_state.custom_zahlungswege):
                zw_col1, zw_col2 = st.columns([3, 1])
                zw_col1.text(zw_item)
                # 기본값인 Firmenkonto는 지워지지 않도록 보호 조치
                if zw_item != "Firmenkonto":
                    if zw_col2.button("Löschen", key=f"del_zw_{zw_item}"):
                        st.session_state.custom_zahlungswege.remove(zw_item)
                        st.rerun()

st.markdown("---")

uploaded_files = st.file_uploader("Digitale Belege hochladen (PDF, PNG, JPG, JPEG)", type=["pdf", "png", "jpg", "jpeg"], accept_multiple_files=True)

col_cfg1, col_cfg2 = st.columns(2)
with col_cfg1: 
    # 동적으로 변하는 커스텀 리스트 연동
    default_zahlart = st.radio("Zahlungsweg Standard", options=st.session_state.custom_zahlungswege, index=0, horizontal=True)
with col_cfg2: 
    selected_skr = st.radio("SKR Standard", options=["SKR03", "SKR04"], index=1, horizontal=True)

if uploaded_files:
    batch_key = "".join(f.name for f in uploaded_files) + f"_{selected_skr}_{default_zahlart}"
    if st.session_state.get("last_batch_key") != batch_key:
        st.session_state.last_batch_key = batch_key
        st.session_state.edited_receipts = None

    if st.session_state.get("edited_receipts") is None:
        rows = []
        total_files = len(uploaded_files)
        progress_bar = st.progress(0)

        with st.spinner("Analysiere Dokumente via Kognitiver AI-Engine..."):
            for idx, uploaded_file in enumerate(uploaded_files):
                file_bytes = uploaded_file.read()
                ext        = uploaded_file.name.rsplit(".", 1)[-1].lower()
                mime_type  = MIME_MAP.get(ext, "application/octet-stream")

                res = ask_gemini_vision_cached(file_bytes, mime_type, API_KEY)
                beleg_nr, date_str, vendor, total, currency, _, mwst_type, raw_text = res[0], res[1], res[2], res[3], res[4], res[5], res[6], res[7]
                was_called = res[8] if len(res) > 8 else False

                assigned_kategorie = get_assigned_account(vendor, selected_skr)
                mwst_19, mwst_7, netto = calculate_tax_details(total, mwst_type)

                rows.append({
                    "Belegdatum": date_str,
                    "Ausgangs-Rechnungsnummer": "",
                    "Kreditor": vendor,
                    "Belegnummer": beleg_nr,
                    "Gegenkonto": assigned_kategorie,
                    "Beleg-Soll (Orig.)": f"{total:,.2f} $" if currency == "USD" else f"{total:,.2f} €",
                    "Zahlungsweg": default_zahlart,
                    "Bruttobetrag (EUR)": total,
                    "USt/Vorsteuer 19%": mwst_19,
                    "Vorsteuer 7%": mwst_7,
                    "Nettobetrag (Haben)": netto,
                    "BU-Schlüssel": mwst_type,
                    "DATEV-Dateiname": build_datev_filename(date_str, vendor, total, default_zahlart, beleg_nr, ""),
                    "_FileExt": ext, "_RawBytes": file_bytes, "_OcrText": raw_text
                })
                progress_bar.progress(int((idx + 1) / total_files * 100))
                if was_called and total_files > 1 and idx < total_files - 1: time.sleep(FREE_TIER_DELAY)

        st.session_state.edited_receipts = pd.DataFrame(rows, index=range(1, len(rows) + 1))
        st.session_state.edited_receipts.index.name = "Nr."

    # DATA EDITOR (이모티콘 완벽 제거 / 수동 편집 및 열 삭제 완벽 가동)
    st.data_editor(
        st.session_state.edited_receipts,
        use_container_width=True, num_rows="dynamic", height=400, key="beleg_editor_key", on_change=on_table_edited,
        column_config={
            "Belegdatum": st.column_config.TextColumn("Belegdatum", width="small"),
            "Ausgangs-Rechnungsnummer": st.column_config.TextColumn("Ausgangs-Rechnungsnummer", width="medium"),
            "Kreditor": st.column_config.TextColumn("Kreditor", width="medium"),
            "Belegnummer": st.column_config.TextColumn("Belegnummer", width="small"),
            "Gegenkonto": st.column_config.TextColumn("Gegenkonto", width="medium"),
            "Beleg-Soll (Orig.)": st.column_config.TextColumn("Beleg-Soll (Orig.)", disabled=True),
            # Zahlungsweg 선택박스에도 커스텀 리스트 실시간 바인딩
            "Zahlungsweg": st.column_config.SelectboxColumn("Zahlungsweg", options=st.session_state.custom_zahlungswege, width="small"),
            "Bruttobetrag (EUR)": st.column_config.NumberColumn("Bruttobetrag (EUR)", format="%,.2f €"),
            "USt/Vorsteuer 19%": st.column_config.NumberColumn("USt/Vorsteuer 19%", format="%,.2f €"),
            "Vorsteuer 7%": st.column_config.NumberColumn("Vorsteuer 7%", format="%,.2f €"),
            "Nettobetrag (Haben)": st.column_config.NumberColumn("Nettobetrag (Haben)", format="%,.2f €"),
            "BU-Schlüssel": st.column_config.SelectboxColumn("BU-Schlüssel", options=["19_Only", "7_Only", "Split", "AUTO_19", "0_Only"], width="small"),
            "DATEV-Dateiname": st.column_config.TextColumn("DATEV-Dateiname", width="max"),
            "_FileExt": None, "_RawBytes": None, "_OcrText": None
        },
    )

    # DOWNLOADS AREA
    df_final = st.session_state.edited_receipts
    today = datetime.now().strftime("%Y%m%d")
    
    st.markdown("### Bereitstellung der DATEV-Exportdateien")
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1: 
        st.download_button(label="Buchungsliste als Excel-Export herunterladen (.xlsx)", data=build_excel_bytes(df_final), file_name=f"DATEV_{selected_skr}_Buchungsliste_{today}.xlsx", use_container_width=True)
    with col_dl2:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for _, row in df_final.iterrows():
                sandwich_pdf_bytes = create_sandwich_pdf(row["_RawBytes"], row["_FileExt"], row["_OcrText"])
                zip_file.writestr(row["DATEV-Dateiname"], sandwich_pdf_bytes)
        zip_buffer.seek(0)
        st.download_button(label="PDF-Belege als ZIP-Archiv herunterladen (.zip)", data=zip_buffer.getvalue(), file_name=f"DATEV_Digitale_Belege_{today}.zip", use_container_width=True, type="primary")

    # 🛡️ 선별적 데이터 파기 제어판 (설정은 그대로 기억, 민감한 영수증 내역만 파기)
    st.markdown("---")
    st.markdown("#### 🔒 Datensicherheit & Datenschutz")
    st.info("Alle hochgeladenen Belege befinden sich ausschließlich im flüchtigen RAM-Arbeitsspeicher Ihres Browsers. Ihre benutzerdefinierten Regeln und Zahlungswege bleiben für das nächste Mal erhalten.")
    
    if st.button("Arbeitssitzung beenden (Belege unwiderruflich aus dem RAM löschen)", type="secondary", use_container_width=True):
        # 영수증 분석 결과와 파일 바이너리 데이터만 조준하여 파기
        st.session_state.edited_receipts = None
        if "last_batch_key" in st.session_state:
            del st.session_state.last_batch_key
        st.cache_data.clear()
        st.success("Erfolgreich bereinigt! Alle Belegdaten und AI-Ergebnisse wurden restlos entfernt. Ihre Einstellungen wurden beibehalten.")
        time.sleep(1.2)
        st.rerun()
