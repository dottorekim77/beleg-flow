import io
import re
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_ILLEGAL_CHARS = re.compile(r'[\\/*?:"<>|]')

def sanitize_filename(text: str) -> str: 
    return _ILLEGAL_CHARS.sub("", text).strip()

def build_datev_filename(date_str: str, vendor: str, brutto_eur: float, zahlart: str, beleg_nr: str, inv_nr: str) -> str:
    """DATEV 표준 명명 규칙에 의거한 PDF 파일명 동적 생성"""
    if str(zahlart).lower() == "firmenkonto": z_code = "B"
    elif str(zahlart).lower() == "kreditkarte": z_code = "C"
    elif str(zahlart).lower() == "bar": z_code = "BAR"
    else: z_code = sanitize_filename(str(zahlart)).replace(" ", "").upper()[:3]

    v_clean = sanitize_filename(vendor).replace(" ", "")[:10]
    b_suffix = f"_{sanitize_filename(beleg_nr)[:12]}" if beleg_nr and beleg_nr.lower() not in ("", "none") else ""
    inv_suffix = f"-I{sanitize_filename(inv_nr)[:8]}" if inv_nr and inv_nr.lower() not in ("", "none") else ""
    return f"{date_str.replace('-', '')}_{v_clean}_{brutto_eur:.2f}EUR_{z_code}{b_suffix}{inv_suffix}.pdf"

def build_excel_bytes(df: pd.DataFrame) -> bytes:
    """openpyxl 엔진 기반 복식부기 감사용 정형화 엑셀 시트 생성"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_clean = df.drop(columns=["_FileExt", "_RawBytes", "_OcrText"], errors="ignore")
        df_clean.to_excel(writer, sheet_name="DATEV_Export", index=True)
        ws = writer.sheets["DATEV_Export"]
        
        HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
        HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        thin = Side(style="thin", color="D9D9D9")
        border_style = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]: 
            cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, border_style
            
        for row in ws.iter_rows(min_row=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = border_style
                if col_idx in (5, 6, 7, 8): cell.number_format = '#,##0.00" €"'
                elif col_idx == 4: cell.alignment = Alignment(horizontal="right")

        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    str_len = sum(2 if ord(char) > 128 else 1 for char in str(cell.value))
                    if str_len > max_len: max_len = str_len
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 5, 16)
    return buf.getvalue()
